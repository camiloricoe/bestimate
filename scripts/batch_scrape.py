"""Batch scrape properties from CSV or Excel.

Usage:
    python scripts/batch_scrape.py "100 props test 1.xlsx" -o results.csv
    python scripts/batch_scrape.py input.csv -o results.csv --limit 50

Preserves all original columns (PropertyId, APN, etc.) and adds Zillow data.
"""

import csv
import logging
import sys
import time
from pathlib import Path

import click

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.scraper.client import ZillowScraper

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# Zillow data columns to add
ZILLOW_COLS = [
    "zpid", "zestimate", "price", "beds", "baths", "sqft",
    "lot_size_sqft", "year_built", "property_type", "zillow_status",
]


def load_excel(path: str) -> list[dict]:
    """Load from .xlsx, normalize column names."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({h: (str(v).strip() if v is not None else "") for h, v in zip(headers, row)})
    wb.close()
    return rows


def load_csv(path: str) -> list[dict]:
    """Load from CSV (auto-detect delimiter)."""
    with open(path) as f:
        sample = f.read(2000)
    delimiter = "\t" if "\t" in sample else ","
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        return [{k.strip(): v.strip() for k, v in row.items()} for row in reader]


def find_column(row: dict, candidates: list[str]) -> str:
    """Find a column value by trying multiple possible names."""
    for c in candidates:
        for key in row:
            if key.lower().replace(" ", "").replace("_", "") == c.lower().replace(" ", "").replace("_", ""):
                if row[key]:
                    return row[key]
    return ""


def write_results(path: str, results: list[dict]):
    if not results:
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=results[0].keys())
        writer.writeheader()
        writer.writerows(results)


@click.command()
@click.argument("input_file", type=click.Path(exists=True))
@click.option("--output", "-o", default=None, help="Output CSV path")
@click.option("--delay", "-d", default=2.0, help="Delay between requests (seconds)")
@click.option("--limit", "-l", default=0, help="Max properties (0=all)")
def main(input_file: str, output: str | None, delay: float, limit: int):
    """Batch scrape Zillow properties."""

    if output is None:
        stem = Path(input_file).stem
        output = str(Path(input_file).parent / f"{stem}_results.csv")

    # Load input
    ext = Path(input_file).suffix.lower()
    if ext in (".xlsx", ".xls"):
        rows = load_excel(input_file)
    else:
        rows = load_csv(input_file)

    total = len(rows) if limit == 0 else min(limit, len(rows))
    logger.info("Loaded %d rows, scraping %d", len(rows), total)

    if rows:
        logger.info("Columns: %s", list(rows[0].keys()))

    # Scrape
    scraper = ZillowScraper()
    results = []
    start_time = time.time()

    for i, row in enumerate(rows[:total]):
        address = find_column(row, ["PropertyAddress", "Property Address", "address", "addr"])
        city = find_column(row, ["CITY", "city"])
        state = find_column(row, ["STATE", "state"])
        zip_code = find_column(row, ["PropertyZip", "Property Zip", "zip", "zipcode", "zip_code"])

        if not address:
            continue

        t0 = time.time()
        prop = scraper.search_property(address, city, state, zip_code)
        elapsed = time.time() - t0

        # Build result: all original columns + zillow data
        result = dict(row)  # Preserve ALL original columns
        result["zpid"] = prop.zpid if prop else ""
        result["zestimate"] = prop.zestimate if prop else ""
        result["price"] = prop.price if prop else ""
        result["beds"] = prop.beds if prop else ""
        result["baths"] = prop.baths if prop else ""
        result["sqft"] = prop.sqft if prop else ""
        result["lot_size_sqft"] = prop.lot_size_sqft if prop else ""
        result["year_built"] = prop.year_built if prop else ""
        result["property_type"] = prop.property_type if prop else ""
        result["zillow_status"] = "ok" if prop else "not_found"
        results.append(result)

        zest = f"${prop.zestimate:,.0f}" if prop and prop.zestimate else "N/A"
        icon = "+" if prop else "-"
        logger.info("[%d/%d %s %.1fs] %s, %s -> %s", i + 1, total, icon, elapsed, address, city, zest)

        # Save every 25
        if (i + 1) % 25 == 0:
            write_results(output, results)
            elapsed_total = time.time() - start_time
            rate = (i + 1) / elapsed_total * 60
            eta = (total - i - 1) / ((i + 1) / elapsed_total) if i > 0 else 0
            logger.info(
                "=== %d/%d | %.0f/min | ETA %.0fm | %s ===",
                i + 1, total, rate, eta / 60, scraper.stats,
            )

        time.sleep(delay)

    # Final save
    write_results(output, results)

    elapsed_total = time.time() - start_time
    stats = scraper.stats
    logger.info("=" * 60)
    logger.info("DONE: %d/%d scraped in %.0fs (%.1fs avg)", stats["scraped"], total, elapsed_total, elapsed_total / total if total else 0)
    logger.info("Stats: %s", stats)
    logger.info("Output: %s", output)
    scraper.close()


if __name__ == "__main__":
    main()
