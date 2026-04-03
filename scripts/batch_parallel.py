"""Parallel batch scraper with retry queue.

Usage:
    python scripts/batch_parallel.py "100 props test 1.xlsx" -o results.csv -w 50
    python scripts/batch_parallel.py input.csv -o results.csv --workers 100 --limit 1000

Each worker has its own sticky proxy session + cookies.
Failed properties are retried automatically.
"""

import csv
import logging
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from queue import Queue

import click

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.scraper.client import ZillowScraper

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# Thread-safe counters
_lock = threading.Lock()
_scraped = 0
_failed = 0
_no_zpid = 0
_blocked = 0
_results: list[dict] = []


@dataclass
class PropertyTask:
    index: int
    original_row: dict
    address: str
    city: str
    state: str
    zip_code: str
    attempt: int = 1


def find_column(row: dict, candidates: list[str]) -> str:
    for c in candidates:
        for key in row:
            if key.lower().replace(" ", "").replace("_", "") == c.lower().replace(" ", "").replace("_", ""):
                if row[key]:
                    return row[key]
    return ""


def load_input(path: str) -> list[dict]:
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({h: (str(v).strip() if v is not None else "") for h, v in zip(headers, row)})
        wb.close()
        return rows
    else:
        with open(path) as f:
            sample = f.read(2000)
        delimiter = "\t" if "\t" in sample else ","
        with open(path, newline="", encoding="utf-8-sig") as f:
            return [{k.strip(): v.strip() for k, v in row.items()} for row in csv.DictReader(f, delimiter=delimiter)]


def scrape_one(task: PropertyTask, scraper: ZillowScraper) -> dict:
    """Scrape a single property. Returns result dict."""
    global _scraped, _failed, _no_zpid, _blocked

    prop = scraper.search_property(task.address, task.city, task.state, task.zip_code)

    result = dict(task.original_row)

    # Track counts before to detect what happened
    no_zpid_before = scraper._total_no_zpid
    failed_before = scraper._total_failed

    if prop:
        result.update({
            "zpid": prop.zpid,
            "zestimate": prop.zestimate or "",
            "price": prop.price or "",
            "beds": prop.beds if prop.beds is not None else "",
            "baths": prop.baths if prop.baths is not None else "",
            "sqft": prop.sqft if prop.sqft is not None else "",
            "lot_size_sqft": prop.lot_size_sqft if prop.lot_size_sqft is not None else "",
            "year_built": prop.year_built if prop.year_built is not None else "",
            "property_type": prop.property_type or "",
            "zillow_status": "ok",
        })
        with _lock:
            _scraped += 1
    else:
        # Distinguish: autocomplete miss vs PX block
        was_no_zpid = scraper._total_no_zpid > no_zpid_before
        status = "not_found" if was_no_zpid else "blocked"
        result.update({
            "zpid": "", "zestimate": "", "price": "", "beds": "", "baths": "",
            "sqft": "", "lot_size_sqft": "", "year_built": "", "property_type": "",
            "zillow_status": status,
        })
        with _lock:
            if was_no_zpid:
                _no_zpid += 1
            # blocked count handled by retry logic

    return result


def worker_fn(tasks: list[PropertyTask], worker_id: int, retry_queue: Queue, delay: float):
    """Worker thread: own scraper instance, processes assigned tasks."""
    scraper = ZillowScraper()

    for task in tasks:
        try:
            result = scrape_one(task, scraper)

            # Only store result if it's the final attempt or success
            if result["zillow_status"] == "blocked" and task.attempt < 3:
                task.attempt += 1
                retry_queue.put(task)
                # Don't store blocked results yet - wait for retry
            else:
                with _lock:
                    _results.append(result)

            time.sleep(delay)

        except Exception as e:
            logger.error("Worker %d error on %s: %s", worker_id, task.address, e)
            if task.attempt < 3:
                task.attempt += 1
                retry_queue.put(task)
            else:
                with _lock:
                    result = dict(task.original_row)
                    result.update({
                        "zpid": "", "zestimate": "", "price": "", "beds": "", "baths": "",
                        "sqft": "", "lot_size_sqft": "", "year_built": "", "property_type": "",
                        "zillow_status": "error",
                    })
                    _results.append(result)

    scraper.close()


def write_results(path: str):
    with _lock:
        if not _results:
            return
        # Sort by original index to maintain order
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=_results[0].keys())
            writer.writeheader()
            writer.writerows(_results)


@click.command()
@click.argument("input_file", type=click.Path(exists=True))
@click.option("--output", "-o", default=None, help="Output CSV path")
@click.option("--workers", "-w", default=20, help="Number of parallel workers")
@click.option("--delay", "-d", default=1.0, help="Delay between requests per worker (seconds)")
@click.option("--limit", "-l", default=0, help="Max properties (0=all)")
@click.option("--skip", "-s", default=0, help="Skip first N rows")
@click.option("--max-retries", "-r", default=2, help="Max retry rounds for blocked properties")
def main(input_file: str, output: str | None, workers: int, delay: float, limit: int, skip: int, max_retries: int):
    """Parallel batch scrape with retry queue."""
    global _scraped, _failed, _no_zpid, _blocked, _results

    if output is None:
        stem = Path(input_file).stem
        output = str(Path(input_file).parent / f"{stem}_results.csv")

    rows = load_input(input_file)
    if skip > 0:
        rows = rows[skip:]
        logger.info("Skipped first %d rows", skip)
    total = len(rows) if limit == 0 else min(limit, len(rows))
    logger.info("Loaded %d rows, scraping %d with %d workers", len(rows), total, workers)

    # Build tasks
    tasks = []
    for i, row in enumerate(rows[:total]):
        address = find_column(row, ["PropertyAddress", "Property Address", "address"])
        city = find_column(row, ["CITY", "city"])
        state = find_column(row, ["STATE", "state"])
        zip_code = find_column(row, ["PropertyZip", "Property Zip", "zip", "zipcode"])
        if address:
            tasks.append(PropertyTask(i, row, address, city, state, zip_code))

    logger.info("Created %d tasks", len(tasks))

    # Distribute tasks across workers
    retry_queue: Queue = Queue()
    start_time = time.time()

    # Progress monitor
    def monitor():
        while not _done.is_set():
            elapsed = time.time() - start_time
            rate = (_scraped + _failed + _no_zpid) / elapsed * 60 if elapsed > 0 else 0
            total_done = _scraped + _failed + _no_zpid
            remaining = len(tasks) - total_done
            eta = remaining / (rate / 60) if rate > 0 else 0
            logger.info(
                "Progress: %d/%d | OK: %d | NoZpid: %d | Blocked: %d | %.0f/min | ETA: %.0fm",
                total_done, len(tasks), _scraped, _no_zpid, _failed, rate, eta / 60,
            )
            write_results(output)
            _done.wait(15)

    _done = threading.Event()
    monitor_thread = threading.Thread(target=monitor, daemon=True)
    monitor_thread.start()

    # Split tasks into chunks for each worker
    chunk_size = max(1, len(tasks) // workers)
    chunks = [tasks[i:i + chunk_size] for i in range(0, len(tasks), chunk_size)]

    logger.info("Starting %d workers with ~%d tasks each...", len(chunks), chunk_size)

    # Run workers
    with ThreadPoolExecutor(max_workers=min(workers, len(chunks))) as executor:
        futures = []
        for wid, chunk in enumerate(chunks):
            futures.append(executor.submit(worker_fn, chunk, wid, retry_queue, delay))

        # Wait for all workers
        for f in as_completed(futures):
            try:
                f.result()
            except Exception as e:
                logger.error("Worker crashed: %s", e)

    # Retry rounds (up to 2 rounds for blocked properties)
    for retry_round in range(1, max_retries + 1):
        retry_tasks = []
        while not retry_queue.empty():
            retry_tasks.append(retry_queue.get())

        if not retry_tasks:
            break

        logger.info("=== Retry round %d: %d blocked properties ===", retry_round, len(retry_tasks))
        next_retry_queue: Queue = Queue()
        retry_chunk_size = max(1, len(retry_tasks) // min(workers, len(retry_tasks)))
        retry_chunks = [retry_tasks[i:i + retry_chunk_size] for i in range(0, len(retry_tasks), retry_chunk_size)]

        with ThreadPoolExecutor(max_workers=min(workers // 2 or 1, len(retry_chunks))) as executor:
            futures = [executor.submit(worker_fn, chunk, wid, next_retry_queue, delay * 1.5) for wid, chunk in enumerate(retry_chunks)]
            for f in as_completed(futures):
                try:
                    f.result()
                except Exception as e:
                    logger.error("Retry worker crashed: %s", e)

        # Remaining failed go to final results
        retry_queue = next_retry_queue

    # Any still in queue after all retries -> mark as failed
    while not retry_queue.empty():
        task = retry_queue.get()
        with _lock:
            result = dict(task.original_row)
            result.update({
                "zpid": "", "zestimate": "", "price": "", "beds": "", "baths": "",
                "sqft": "", "lot_size_sqft": "", "year_built": "", "property_type": "",
                "zillow_status": "blocked_final",
            })
            _results.append(result)
            _failed += 1

    _done.set()

    # Final save
    write_results(output)

    elapsed = time.time() - start_time
    rate = (_scraped + _failed + _no_zpid) / elapsed * 60
    logger.info("=" * 60)
    logger.info("DONE in %.0fs (%.0f/min)", elapsed, rate)
    logger.info("Scraped: %d | No ZPID: %d | Failed: %d | Total: %d/%d",
                _scraped, _no_zpid, _failed, _scraped + _no_zpid + _failed, len(tasks))
    logger.info("Success rate: %.0f%%", _scraped / len(tasks) * 100 if tasks else 0)
    logger.info("Output: %s", output)


if __name__ == "__main__":
    main()
